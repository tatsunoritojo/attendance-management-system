"""
Local CSV data handling for Attendance Management System v3.4
This module replaces the Google Sheets integration with local CSV file operations,
including persistent storage for attendance records.
"""

import csv
import logging
from datetime import datetime
from functools import lru_cache
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from attendance_app.path_manager import get_asset_path, get_output_dir

logger = logging.getLogger(__name__)

# Define the path for the persistent attendance history CSV
ATTENDANCE_HISTORY_FILE = get_output_dir() / "attendance_history.csv"

class CsvDataError(Exception):
    """Custom exception for CSV data handling errors."""
    pass

def _initialize_attendance_history():
    """Ensures the attendance history CSV file exists with the correct header."""
    if not ATTENDANCE_HISTORY_FILE.exists():
        ATTENDANCE_HISTORY_FILE.parent.mkdir(parents=True, exist_ok=True)
        with open(ATTENDANCE_HISTORY_FILE, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerow(["Entry_Time", "StudentID", "Name", "Mood", "Sleep_Satisfaction", "Purpose", "Exit_Time"])

# Initialize the history file on module load
_initialize_attendance_history()

@lru_cache()
def _read_student_data_from_excel() -> List[Dict[str, str]]:
    """Reads student data from the local Sample_Data.xlsx file (StudentID_StudentName sheet)."""
    # Try xlsx first, then fall back to xlsm
    excel_file_path = get_asset_path('Sample_Data.xlsx')
    if not excel_file_path.exists():
        excel_file_path = get_asset_path('Sample_Data.xlsm')
    if not excel_file_path.exists():
        raise CsvDataError(f"Student data file not found: {excel_file_path}")

    student_list = []
    try:
        # openpyxl を使用してExcelファイルを読み込む
        from openpyxl import load_workbook
        workbook = load_workbook(excel_file_path, data_only=True)
        sheet = workbook['StudentID_StudentName'] # StudentID_StudentName シートを指定

        # ヘッダー行をスキップし、データを読み込む
        headers = [cell.value for cell in sheet[1]] # 1行目をヘッダーとして取得
        
        # StudentIDとStudentNameの列インデックスを特定
        student_id_col_idx = -1
        student_name_col_idx = -1
        for i, header in enumerate(headers):
            if header == 'StudentID':
                student_id_col_idx = i
            elif header == 'StudentName':
                student_name_col_idx = i

        if student_id_col_idx == -1 or student_name_col_idx == -1:
            raise CsvDataError("Required headers 'StudentID' or 'StudentName' not found in StudentID_StudentName sheet.")

        for row_idx in range(2, sheet.max_row + 1): # 2行目からデータを読み込む
            row_data = [cell.value for cell in sheet[row_idx]]
            if len(row_data) > max(student_id_col_idx, student_name_col_idx):
                student_id = str(row_data[student_id_col_idx]).strip()
                student_name = str(row_data[student_name_col_idx]).strip()
                student_list.append({"id": student_id, "name": student_name})
    except ImportError:
        raise CsvDataError("openpyxl library not found. Please install it: pip install openpyxl")
    except KeyError:
        raise CsvDataError("Sheet 'StudentID_StudentName' not found in Sample_Data.xlsm. Please check the sheet name.")
    except Exception as e:
        raise CsvDataError(f"Failed to process student data Excel: {e}")
    return student_list

def get_student_name(student_id: str) -> str:
    """Get student name by ID from the local Excel data."""
    try:
        all_students = _read_student_data_from_excel()
        for student in all_students:
            if student["id"] == str(student_id):
                return student["name"]
        return "Unknown"
    except CsvDataError:
        return "Unknown"

def get_last_record(student_id: str) -> Tuple[Optional[int], Optional[str]]:
    """Gets the last record for a student from the CSV file."""
    try:
        df = pd.read_csv(ATTENDANCE_HISTORY_FILE, dtype=str, encoding='utf-8-sig')
        student_records = df[df['StudentID'] == str(student_id)]
        
        if not student_records.empty:
            last_record = student_records.iloc[-1]
            # If ExitTime is missing or empty, the student is considered "in"
            if pd.isna(last_record.get('Exit_Time')) or not last_record.get('Exit_Time'):
                return int(last_record.name) + 2, None  # Return DataFrame index + 2 to mimic 1-based gspread row
        return None, "dummy_exit_time" # No open entry found
    except FileNotFoundError:
        return None, None
    except Exception as e:
        logger.error(f"Error reading attendance history: {e}")
        return None, None

def append_entry(student_id: str, student_name: str) -> Optional[int]:
    """Appends a new entry to the attendance history CSV."""
    try:
        # Read current data to get the number of existing rows
        try:
            df_before_append = pd.read_csv(ATTENDANCE_HISTORY_FILE, dtype=str, encoding='utf-8-sig')
            num_existing_data_rows = len(df_before_append)
        except FileNotFoundError:
            num_existing_data_rows = 0 # File doesn't exist yet, so 0 data rows

        entry_time = datetime.now().strftime("%Y/%m/%d %H:%M:%S")
        new_record = [entry_time, student_id, student_name, "", "", "", ""]
        
        with open(ATTENDANCE_HISTORY_FILE, 'a', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerow(new_record)
        
        # The new row's 1-based index in the CSV file (including header)
        # If 0 existing data rows, new row is 1st data row, which is line 2. (0 + 2)
        # If 1 existing data row, new row is 2nd data row, which is line 3. (1 + 2)
        return num_existing_data_rows + 2
    except Exception as e:
        logger.error(f"Failed to append entry: {e}")
        return None

def write_response(row: int, col: int, value: str) -> bool:
    """Writes a response (mood, sleep, etc.) to the specified row in the CSV."""
    try:
        df = pd.read_csv(ATTENDANCE_HISTORY_FILE, dtype=str, encoding='utf-8-sig')
        # Convert 1-based Kivy row to 0-based DataFrame index
        df_index = row - 2
        
        logger.debug(f"write_response: row={row}, col={col}, value='{value}'")
        logger.debug(f"write_response: df_index={df_index}")

        if 0 <= df_index < len(df):
            col_name = df.columns[col - 1] # Convert 1-based col to 0-based col index
            logger.debug(f"write_response: col_name='{col_name}'")
            df.loc[df_index, col_name] = value
            df.to_csv(ATTENDANCE_HISTORY_FILE, index=False, encoding='utf-8-sig')
            logger.debug(f"write_response: Successfully wrote to {ATTENDANCE_HISTORY_FILE}")
            return True
        logger.warning(f"write_response: Invalid df_index {df_index} for DataFrame of length {len(df)}")
        return False
    except Exception as e:
        logger.error(f"Failed to write response: {e}")
        return False

def write_exit(row: int) -> bool:
    """Writes the exit time to the specified row in the CSV."""
    try:
        exit_time = datetime.now().strftime("%Y/%m/%d %H:%M:%S")
        return write_response(row, 7, exit_time) # Column G (Exit_Time) is the 7th column
    except Exception as e:
        logger.error(f"Failed to write exit time: {e}")
        return False

def get_student_list_for_printing() -> List[Dict[str, str]]:
    """Gets the list of all students from the Excel for printing purposes."""
    try:
        return _read_student_data_from_excel()
    except CsvDataError as e:
        logger.error(f"Could not get student list for printing: {e}")
        return []

def sync_attendance_to_excel() -> bool:
    """
    attendance_history.csv の出席情報を Sample_Data.xlsm の Attendance_Information シートに同期する。
    """
    try:
        # 1. attendance_history.csv を読み込む
        if not ATTENDANCE_HISTORY_FILE.exists():
            logger.info("attendance_history.csv does not exist. Nothing to sync.")
            return True

        df_csv = pd.read_csv(ATTENDANCE_HISTORY_FILE, encoding='utf-8-sig', dtype=str)
        # CSVのヘッダーをExcelのヘッダーに合わせる
        df_csv = df_csv.rename(columns={
            "EntryTime": "Entry_Time", # 既存のCSVがEntryTimeの場合の対応
            "StudentName": "Name",     # 既存のCSVがStudentNameの場合の対応
            "Sleep": "Sleep_Satisfaction", # 既存のCSVがSleepの場合の対応
            "ExitTime": "Exit_Time"    # 既存のCSVがExitTimeの場合の対応
        })
        # 必要な列のみを選択し、順序を合わせる
        required_cols = ["Entry_Time", "StudentID", "Name", "Mood", "Sleep_Satisfaction", "Purpose", "Exit_Time"]
        df_csv = df_csv[required_cols]

        # 2. Sample_Data.xlsx を読み込む
        excel_file_path = get_asset_path('Sample_Data.xlsx')
        if not excel_file_path.exists():
            excel_file_path = get_asset_path('Sample_Data.xlsm')
        if not excel_file_path.exists():
            logger.error(f"Sample_Data file not found: {excel_file_path}")
            return False

        workbook = load_workbook(excel_file_path)
        sheet_name = 'Attendance_Information'
        
        if sheet_name not in workbook.sheetnames:
            logger.warning(f"Sheet '{sheet_name}' not found in Sample_Data.xlsm. Creating new sheet.")
            sheet = workbook.create_sheet(sheet_name)
            # ヘッダーを書き込む
            sheet.append(required_cols)
            df_excel = pd.DataFrame(columns=required_cols) # 新規作成時は空のDataFrame
        else:
            sheet = workbook[sheet_name]
            # ExcelのデータをDataFrameに変換
            data = sheet.values
            cols = next(data) # ヘッダー行
            df_excel = pd.DataFrame(data, columns=cols)
            # ExcelのヘッダーがCSVのヘッダーと異なる場合のリネーム
            df_excel = df_excel.rename(columns={
                "EntryTime": "Entry_Time",
                "StudentName": "Name",
                "Sleep": "Sleep_Satisfaction",
                "ExitTime": "Exit_Time"
            })
            # 必要な列のみを選択し、順序を合わせる
            df_excel = df_excel[required_cols]


        # 3. 差分を検出してExcelに追記する
        # 既存のExcelデータにないCSVの行を特定
        # 厳密な比較のため、両方のDataFrameを文字列型に変換
        df_csv_str = df_csv.astype(str)
        df_excel_str = df_excel.astype(str)

        # Excelに存在しない行を特定
        # マージして、Excelにのみ存在する行をフィルタリング
        merged_df = pd.merge(df_csv_str, df_excel_str, on=required_cols, how='left', indicator=True)
        new_entries = merged_df[merged_df['_merge'] == 'left_only'].drop(columns=['_merge'])

        if not new_entries.empty:
            logger.info(f"Found {len(new_entries)} new entries to sync to Excel.")
            # 新しいエントリをExcelシートに追記
            for r_idx, row_data in enumerate(dataframe_to_rows(new_entries, index=False, header=False)):
                sheet.append(row_data)
            workbook.save(excel_file_path)
            logger.info(f"Successfully synced {len(new_entries)} entries to {excel_file_path}")
            return True
        else:
            logger.info("No new entries to sync to Excel.")
            return True

    except Exception as e:
        logger.error(f"Error during Excel synchronization: {e}")
        return False