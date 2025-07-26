"""
学生データ管理モジュール
StudentListシートとStudentID_StudentNameシートの統合管理
"""

import logging
from datetime import datetime
from typing import Dict, List, Optional, Tuple
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from attendance_app.path_manager import get_asset_path
from attendance_app.spreadsheet import _read_student_data_from_excel

logger = logging.getLogger(__name__)

class StudentDataManager:
    """学生データの統合管理クラス"""
    
    def __init__(self):
        self.excel_file_path = self._get_excel_file_path()
    
    def _get_excel_file_path(self):
        """Excelファイルのパスを取得"""
        excel_file_path = get_asset_path('Sample_Data.xlsx')
        if not excel_file_path.exists():
            excel_file_path = get_asset_path('Sample_Data.xlsm')
        return excel_file_path
    
    def generate_student_id(self) -> str:
        """
        学籍番号を自動生成
        形式: 25D#### (現在の年の下2桁 + D + 4桁連番)
        """
        try:
            # 現在の年の下2桁を取得
            current_year = datetime.now().year
            year_suffix = str(current_year)[-2:]
            
            # 既存の学生IDから最大番号を取得
            existing_students = _read_student_data_from_excel()
            max_number = 0
            
            prefix = f"{year_suffix}D"
            
            for student in existing_students:
                student_id = student['id']
                if student_id.startswith(prefix) and len(student_id) == 7:
                    try:
                        number_part = int(student_id[3:])  # 25D0019 -> 19
                        max_number = max(max_number, number_part)
                    except ValueError:
                        continue
            
            # 次の番号を生成（8間隔で増加するパターンを維持）
            if max_number == 0:
                next_number = 19  # 最初の番号
            else:
                next_number = max_number + 8
            
            # 4桁にゼロパディング
            new_student_id = f"{prefix}{next_number:04d}"
            
            logger.info(f"Generated new student ID: {new_student_id}")
            return new_student_id
            
        except Exception as e:
            logger.error(f"Error generating student ID: {e}")
            # フォールバック: タイムスタンプベース
            timestamp = datetime.now().strftime("%m%d")
            return f"25D{timestamp}"
    
    def create_student_list_sheet_if_not_exists(self):
        """StudentsListシートが存在しない場合は作成"""
        try:
            workbook = load_workbook(self.excel_file_path)
            
            if 'StudentsList' not in workbook.sheetnames:
                logger.info("Creating StudentsList sheet")
                sheet = workbook.create_sheet('StudentsList')
                
                # ヘッダー行を作成
                headers = [
                    "登録日",        # A列
                    "生徒氏名",      # B列
                    "保護者氏名",    # C列
                    "保護者連絡先",  # D列
                    "学校名",        # E列
                    "",              # F列（空）
                    "",              # G列（空）
                    "生年月日"       # H列
                ]
                
                for col, header in enumerate(headers, 1):
                    sheet.cell(row=1, column=col, value=header)
                
                workbook.save(self.excel_file_path)
                logger.info("StudentsList sheet created successfully")
            
        except Exception as e:
            logger.error(f"Error creating StudentsList sheet: {e}")
            raise
    
    def add_student_to_student_list(self, student_data: Dict[str, str]) -> bool:
        """
        StudentsListシートに学生データを追加
        
        Args:
            student_data: {
                'student_name': str,
                'guardian_name': str,
                'guardian_contact': str,
                'school_name': str,
                'birth_date': str
            }
        
        Returns:
            bool: 成功時True
        """
        try:
            self.create_student_list_sheet_if_not_exists()
            
            workbook = load_workbook(self.excel_file_path)
            sheet = workbook['StudentsList']
            
            # 次の空行を見つける
            next_row = sheet.max_row + 1
            
            # 登録日（現在の日時）
            registration_date = datetime.now().strftime("%Y/%m/%d %H:%M:%S")
            
            # データを各列に書き込み
            sheet.cell(row=next_row, column=1, value=registration_date)      # A: 登録日
            sheet.cell(row=next_row, column=2, value=student_data['student_name'])    # B: 生徒氏名
            sheet.cell(row=next_row, column=3, value=student_data['guardian_name'])   # C: 保護者氏名
            sheet.cell(row=next_row, column=4, value=student_data['guardian_contact']) # D: 保護者連絡先
            sheet.cell(row=next_row, column=5, value=student_data['school_name'])     # E: 学校名
            # F, G列は空
            sheet.cell(row=next_row, column=8, value=student_data['birth_date'])      # H: 生年月日
            
            workbook.save(self.excel_file_path)
            logger.info(f"Student data added to StudentsList sheet: {student_data['student_name']}")
            return True
            
        except Exception as e:
            logger.error(f"Error adding student to StudentsList: {e}")
            return False
    
    def add_student_to_student_id_name(self, student_id: str, student_name: str) -> bool:
        """
        StudentID_StudentNameシートに学生IDと名前を追加
        
        Args:
            student_id: 学籍番号
            student_name: 学生名
            
        Returns:
            bool: 成功時True
        """
        try:
            workbook = load_workbook(self.excel_file_path)
            
            if 'StudentID_StudentName' not in workbook.sheetnames:
                logger.info("Creating StudentID_StudentName sheet")
                sheet = workbook.create_sheet('StudentID_StudentName')
                # ヘッダー行を作成
                sheet.cell(row=1, column=1, value='StudentID')
                sheet.cell(row=1, column=2, value='StudentName')
            else:
                sheet = workbook['StudentID_StudentName']
            
            # 次の空行を見つける
            next_row = sheet.max_row + 1
            
            # データを書き込み
            sheet.cell(row=next_row, column=1, value=student_id)
            sheet.cell(row=next_row, column=2, value=student_name)
            
            workbook.save(self.excel_file_path)
            logger.info(f"Student ID added to StudentID_StudentName sheet: {student_id} - {student_name}")
            return True
            
        except Exception as e:
            logger.error(f"Error adding student to StudentID_StudentName sheet: {e}")
            return False
    
    def register_new_student(self, student_data: Dict[str, str]) -> Tuple[bool, str, str]:
        """
        新しい学生を完全登録（StudentListとStudentID_StudentNameの両方に追加）
        
        Args:
            student_data: 学生データ辞書
            
        Returns:
            Tuple[bool, str, str]: (成功フラグ, 学籍番号, エラーメッセージ)
        """
        try:
            # 学籍番号を生成
            student_id = self.generate_student_id()
            
            # StudentsListシートに追加
            if not self.add_student_to_student_list(student_data):
                return False, "", "StudentsListシートへの書き込みに失敗しました"
            
            # StudentID_StudentNameシートに追加
            if not self.add_student_to_student_id_name(student_id, student_data['student_name']):
                return False, student_id, "StudentID_StudentNameシートへの書き込みに失敗しました"
            
            logger.info(f"Successfully registered new student: {student_data['student_name']} ({student_id})")
            return True, student_id, ""
            
        except Exception as e:
            logger.error(f"Error in complete student registration: {e}")
            return False, "", f"学生登録中にエラーが発生しました: {e}"
    
    def get_student_list_data(self) -> List[Dict[str, str]]:
        """
        StudentsListシートからデータを取得
        
        Returns:
            List[Dict]: 学生データのリスト
        """
        try:
            self.create_student_list_sheet_if_not_exists()
            
            workbook = load_workbook(self.excel_file_path, data_only=True)
            sheet = workbook['StudentsList']
            
            students = []
            
            # ヘッダー行をスキップして2行目から読み込み
            for row in range(2, sheet.max_row + 1):
                row_data = []
                for col in range(1, 9):  # A列からH列まで
                    cell_value = sheet.cell(row=row, column=col).value
                    row_data.append(str(cell_value) if cell_value is not None else "")
                
                # 生徒氏名（B列）が空でない場合のみ追加
                if row_data[1].strip():  # B列が空でない
                    student_record = {
                        'registration_date': row_data[0],    # A: 登録日
                        'student_name': row_data[1],         # B: 生徒氏名
                        'guardian_name': row_data[2],        # C: 保護者氏名
                        'guardian_contact': row_data[3],     # D: 保護者連絡先
                        'school_name': row_data[4],          # E: 学校名
                        'birth_date': row_data[7]            # H: 生年月日
                    }
                    students.append(student_record)
            
            return students
            
        except Exception as e:
            logger.error(f"Error reading StudentsList data: {e}")
            return []

# グローバルインスタンス
student_data_manager = StudentDataManager()